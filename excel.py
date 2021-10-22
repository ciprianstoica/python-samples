# pip install openpyxl
from openpyxl import Workbook

wb = Workbook()

ws = wb.active
ws.title = "trilulilu"
ws.cell(1,1, 'alabala')
ws.cell(2,3, 'portocala')

wb.save(filename ='sample_excel.xlsx')